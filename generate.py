"""Build index.html for the Nauta Implementation Dashboard.

Reads data/Nauta_Implementation_Tracker.xlsx, keeps only clients with
owner == 'ss' or 'SS' in the "Resumen Clientes" sheet, and renders a
single-page static HTML using Nauta brand (Manrope + #4A6CF7).
"""
from __future__ import annotations

import datetime
import html
import json
import re
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
XLSX = ROOT / "data" / "Nauta_Implementation_Tracker.xlsx"
OUT = ROOT / "index.html"

EXCLUDE_STAGES = ()  # include all stages — owner filter alone is enough
OWNER_FILTER = {"ss", "SS", "Ss", "sS"}


def to_pct(raw):
    """Parse % avance — accepts '50%', 0.5, 50, '50'."""
    if raw is None:
        return 0
    try:
        s = str(raw).replace("%", "").strip()
        v = float(s)
        return round(v * 100) if v <= 1 else round(v)
    except (ValueError, TypeError):
        return 0


def fallback_pct(stage: str, handover: str) -> int:
    if handover == "Ready to handover":
        return 85
    by_stage = {"5.": 85, "4.": 60, "3.": 40, "2.": 15, "1.": 10}
    for prefix, val in by_stage.items():
        if stage.startswith(prefix):
            return val
    return 10


def clean(s):
    if s is None:
        return ""
    out = str(s).strip()
    if out.startswith("="):  # unresolved formula
        return ""
    return out.replace("\n", " · ").replace('"', "'")


def fmt_due(raw):
    if raw is None:
        return ""
    if isinstance(raw, datetime.datetime):
        return raw.strftime("%Y-%m-%d")
    s = str(raw).strip()
    if s in {"?", "-"}:
        return s
    return s[:20]


def status_class(status: str) -> str:
    s = (status or "").lower()
    if "complet" in s or "done" in s or "ready" in s:
        return "ok"
    if "progreso" in s or "progress" in s:
        return "wip"
    if "bloque" in s or "blocked" in s:
        return "blk"
    return "pen"


def load_clients():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws_clients = wb["Resumen Clientes"]
    hdr = [c.value for c in ws_clients[1]]
    col = {h: i for i, h in enumerate(hdr) if h}

    required = ["Cliente", "Etapa Actual", "Handover", "Status", "% Avance", "owner"]
    missing = [c for c in required if c not in col]
    if missing:
        raise KeyError(f"Missing columns in 'Resumen Clientes': {missing}")

    clients_meta = {}
    for row in ws_clients.iter_rows(min_row=2, values_only=True):
        if not row[col["Cliente"]]:
            continue
        owner = row[col["owner"]]
        if not owner or str(owner).strip() not in OWNER_FILTER:
            continue

        name = str(row[col["Cliente"]]).strip()
        stage = clean(row[col["Etapa Actual"]])
        if not stage or any(stage.startswith(x) for x in EXCLUDE_STAGES):
            continue

        handover = clean(row[col["Handover"]])
        status = clean(row[col["Status"]]) or "En Progreso"
        pct = to_pct(row[col["% Avance"]]) or fallback_pct(stage, handover)

        clients_meta[name] = {
            "n": name,
            "k": re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-"),
            "st": stage,
            "sx": status,
            "p": pct,
            "handover": handover,
            "blocker": clean(row[col.get("Blocker Principal", -1)]) if "Blocker Principal" in col else "",
            "next_action": clean(row[col.get("Siguiente Acción", -1)]) if "Siguiente Acción" in col else "",
            "pain_points": clean(row[col.get("Pain Points", -1)]) if "Pain Points" in col else "",
            "ob_lead": clean(row[col.get("OB Lead", -1)]) if "OB Lead" in col else "",
            "kam": clean(row[col.get("KAM", -1)]) if "KAM" in col else "",
        }

    # Tareas
    ws_tasks = wb["Tareas Detalladas"]
    hdr_t = [c.value for c in ws_tasks[1]]
    colt = {h: i for i, h in enumerate(hdr_t) if h}
    by_client = defaultdict(list)
    for row in ws_tasks.iter_rows(min_row=2, values_only=True):
        c = row[colt["Cliente"]]
        if not c:
            continue
        cname = str(c).strip()
        if cname not in clients_meta:
            continue
        tarea = clean(row[colt.get("Tarea", -1)])
        if not tarea:
            continue
        by_client[cname].append({
            "t": tarea,
            "r": clean(row[colt.get("Responsabilidad", -1)]),
            "d": fmt_due(row[colt.get("Due date", -1)]),
            "o": clean(row[colt.get("Owner", -1)]),
            "s": clean(row[colt.get("Status", -1)]),
            "p": clean(row[colt.get("Prioridad", -1)]),
            "ri": clean(row[colt.get("Riesgos", -1)]),
            "ns": clean(row[colt.get("Next Steps", -1)]),
            "lw": clean(row[colt.get("Last Week", -1)]),
        })

    # Compose final list
    STAGE_ORDER = ["1.", "2.", "3.", "4.", "5.", "6.", "7.", "8."]
    def sk(st):
        return next((i for i, k in enumerate(STAGE_ORDER) if st.startswith(k)), 99)

    clients = []
    for name, meta in clients_meta.items():
        tasks = by_client.get(name, [])
        open_tasks = [t for t in tasks if status_class(t["s"]) not in ("ok",)]
        meta["tasks"] = tasks
        meta["task_count"] = len(tasks)
        meta["open_task_count"] = len(open_tasks)
        meta["risks"] = sorted({t["ri"] for t in tasks if t["ri"]})
        meta["next_steps"] = sorted({t["ns"] for t in tasks if t["ns"]})
        meta["last_week"] = sorted({t["lw"] for t in tasks if t["lw"]})
        clients.append(meta)

    clients.sort(key=lambda c: (sk(c["st"]), c["n"]))
    return clients


def render_html(clients) -> str:
    payload = json.dumps(clients, ensure_ascii=False)
    updated = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    n_clients = len(clients)
    n_tasks = sum(c["task_count"] for c in clients)
    n_open = sum(c["open_task_count"] for c in clients)

    return f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Nauta · Implementation Tracker</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Manrope:wght@400;500;600;700;800&display=swap" rel="stylesheet">
<style>
:root {{
  --nauta-blue:    #4A6CF7;
  --nauta-blue-d:  #2F4ED4;
  --nauta-dark:    #0F1B3D;
  --nauta-bg:      #F5F7FB;
  --nauta-card:    #FFFFFF;
  --nauta-border:  #E2E7F1;
  --nauta-text:    #1A2540;
  --nauta-muted:   #6B7380;
  --ok:    #16A34A;
  --wip:   #4A6CF7;
  --blk:   #DC2626;
  --pen:   #D97706;
}}
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
html, body {{ height: 100%; }}
body {{
  font-family: 'Manrope', system-ui, -apple-system, sans-serif;
  background: var(--nauta-bg);
  color: var(--nauta-text);
  font-size: 14px;
  line-height: 1.5;
}}
a {{ color: var(--nauta-blue); text-decoration: none; }}

.topbar {{
  background: var(--nauta-dark);
  color: white;
  padding: 14px 28px;
  display: flex;
  align-items: center;
  justify-content: space-between;
  box-shadow: 0 1px 6px rgba(15,27,61,0.12);
}}
.brand {{
  display: flex;
  align-items: baseline;
  gap: 10px;
}}
.brand .logo {{
  font-weight: 800;
  font-size: 20px;
  letter-spacing: -0.02em;
  color: white;
}}
.brand .logo::before {{
  content: '';
  display: inline-block;
  width: 10px;
  height: 10px;
  background: var(--nauta-blue);
  border-radius: 2px;
  margin-right: 8px;
  transform: translateY(-1px);
}}
.brand .product {{
  font-weight: 500;
  font-size: 13px;
  color: rgba(255,255,255,0.7);
  letter-spacing: 0.04em;
  text-transform: uppercase;
}}
.topbar-stats {{
  display: flex;
  gap: 24px;
  font-size: 12px;
}}
.topbar-stats span b {{ color: white; font-weight: 700; }}
.topbar-stats span {{ color: rgba(255,255,255,0.65); }}

.layout {{
  display: grid;
  grid-template-columns: 280px 1fr;
  height: calc(100vh - 56px);
}}

.sidebar {{
  background: var(--nauta-card);
  border-right: 1px solid var(--nauta-border);
  overflow-y: auto;
  padding: 16px 0;
}}
.sidebar h3 {{
  font-size: 11px;
  text-transform: uppercase;
  letter-spacing: 0.08em;
  color: var(--nauta-muted);
  padding: 0 20px 8px;
  font-weight: 700;
}}
.client-item {{
  padding: 12px 20px;
  border-left: 3px solid transparent;
  cursor: pointer;
  transition: background 0.15s;
}}
.client-item:hover {{ background: var(--nauta-bg); }}
.client-item.active {{
  background: var(--nauta-bg);
  border-left-color: var(--nauta-blue);
}}
.client-item .name {{
  font-weight: 600;
  font-size: 13px;
  color: var(--nauta-text);
  margin-bottom: 4px;
}}
.client-item .meta {{
  font-size: 11px;
  color: var(--nauta-muted);
  display: flex;
  justify-content: space-between;
  align-items: center;
}}
.bar {{
  width: 100%;
  height: 4px;
  background: var(--nauta-border);
  border-radius: 99px;
  overflow: hidden;
  margin-top: 6px;
}}
.bar > div {{ height: 100%; background: var(--nauta-blue); }}

.main {{
  overflow-y: auto;
  padding: 28px 36px;
}}
.client-header {{
  display: flex;
  justify-content: space-between;
  align-items: flex-start;
  margin-bottom: 24px;
  padding-bottom: 20px;
  border-bottom: 1px solid var(--nauta-border);
}}
.client-header h1 {{
  font-size: 26px;
  font-weight: 800;
  letter-spacing: -0.02em;
  margin-bottom: 6px;
}}
.client-header .sub {{
  color: var(--nauta-muted);
  font-size: 13px;
}}
.client-header .progress {{
  text-align: right;
  min-width: 200px;
}}
.client-header .progress .pct {{
  font-size: 32px;
  font-weight: 800;
  color: var(--nauta-blue);
}}
.client-header .progress .lbl {{
  font-size: 11px;
  color: var(--nauta-muted);
  text-transform: uppercase;
  letter-spacing: 0.06em;
}}

.cards {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(220px, 1fr)); gap: 16px; margin-bottom: 24px; }}
.card {{
  background: var(--nauta-card);
  border: 1px solid var(--nauta-border);
  border-radius: 12px;
  padding: 16px;
}}
.card h4 {{
  font-size: 11px;
  text-transform: uppercase;
  letter-spacing: 0.06em;
  color: var(--nauta-muted);
  margin-bottom: 8px;
  font-weight: 700;
}}
.card .val {{ font-size: 16px; font-weight: 600; color: var(--nauta-text); }}
.card .val.muted {{ color: var(--nauta-muted); font-weight: 400; font-style: italic; }}

.section {{
  background: var(--nauta-card);
  border: 1px solid var(--nauta-border);
  border-radius: 12px;
  margin-bottom: 16px;
  overflow: hidden;
}}
.section header {{
  padding: 14px 18px;
  background: var(--nauta-card);
  border-bottom: 1px solid var(--nauta-border);
  display: flex;
  align-items: center;
  justify-content: space-between;
}}
.section header h3 {{ font-size: 14px; font-weight: 700; }}
.section header .count {{
  background: var(--nauta-bg);
  border-radius: 99px;
  padding: 2px 10px;
  font-size: 12px;
  font-weight: 600;
  color: var(--nauta-muted);
}}
.section .body {{ padding: 4px 0; }}
.section ul {{ list-style: none; }}
.section ul li {{
  padding: 10px 18px;
  border-bottom: 1px solid var(--nauta-bg);
  font-size: 13px;
}}
.section ul li:last-child {{ border-bottom: none; }}

table.tasks {{
  width: 100%;
  border-collapse: collapse;
}}
table.tasks thead th {{
  text-align: left;
  padding: 10px 14px;
  font-size: 11px;
  text-transform: uppercase;
  letter-spacing: 0.06em;
  color: var(--nauta-muted);
  background: var(--nauta-bg);
  font-weight: 700;
  border-bottom: 1px solid var(--nauta-border);
}}
table.tasks tbody td {{
  padding: 10px 14px;
  font-size: 13px;
  border-bottom: 1px solid var(--nauta-bg);
  vertical-align: top;
}}
table.tasks tbody tr:last-child td {{ border-bottom: none; }}
table.tasks tbody tr:hover {{ background: var(--nauta-bg); }}
.pill {{
  display: inline-block;
  padding: 2px 8px;
  border-radius: 99px;
  font-size: 11px;
  font-weight: 600;
}}
.pill.ok  {{ background: rgba(22,163,74,0.12);  color: var(--ok); }}
.pill.wip {{ background: rgba(74,108,247,0.12); color: var(--wip); }}
.pill.blk {{ background: rgba(220,38,38,0.12);  color: var(--blk); }}
.pill.pen {{ background: rgba(217,119,6,0.12);  color: var(--pen); }}

.empty {{
  padding: 28px;
  color: var(--nauta-muted);
  font-style: italic;
  font-size: 13px;
  text-align: center;
}}

.stage-chip {{
  display: inline-block;
  background: rgba(74,108,247,0.10);
  color: var(--nauta-blue);
  font-weight: 600;
  font-size: 12px;
  padding: 4px 10px;
  border-radius: 99px;
}}
.handover-chip {{
  display: inline-block;
  background: rgba(22,163,74,0.12);
  color: var(--ok);
  font-weight: 600;
  font-size: 12px;
  padding: 4px 10px;
  border-radius: 99px;
  margin-left: 6px;
}}

footer {{
  padding: 16px 28px;
  font-size: 11px;
  color: var(--nauta-muted);
  text-align: center;
  border-top: 1px solid var(--nauta-border);
  background: var(--nauta-card);
}}
</style>
</head>
<body>

<div class="topbar">
  <div class="brand">
    <span class="logo">Nauta</span>
    <span class="product">Implementation Tracker</span>
  </div>
  <div class="topbar-stats">
    <span>Clientes activos · <b>{n_clients}</b></span>
    <span>Tareas totales · <b>{n_tasks}</b></span>
    <span>Tareas abiertas · <b>{n_open}</b></span>
    <span>Actualizado · <b>{updated}</b></span>
  </div>
</div>

<div class="layout">
  <aside class="sidebar">
    <h3>Clientes</h3>
    <div id="client-list"></div>
  </aside>

  <main class="main" id="main"></main>
</div>

<footer>
  Nauta · Implementation Tracker · interno · santiago@getnauta.com
</footer>

<script>
const DATA = {payload};

function statusClass(s) {{
  s = (s || '').toLowerCase();
  if (s.includes('complet') || s.includes('done') || s.includes('ready')) return 'ok';
  if (s.includes('progreso') || s.includes('progress')) return 'wip';
  if (s.includes('bloque') || s.includes('blocked')) return 'blk';
  return 'pen';
}}

function esc(s) {{
  return (s||'').toString()
    .replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');
}}

function renderSidebar(activeK) {{
  const list = document.getElementById('client-list');
  list.innerHTML = DATA.map(c => `
    <div class="client-item ${{c.k === activeK ? 'active' : ''}}" data-k="${{c.k}}">
      <div class="name">${{esc(c.n)}}</div>
      <div class="meta">
        <span>${{esc(c.st)}}</span>
        <span><b>${{c.p}}%</b></span>
      </div>
      <div class="bar"><div style="width:${{c.p}}%"></div></div>
    </div>
  `).join('');
  list.querySelectorAll('.client-item').forEach(el => {{
    el.addEventListener('click', () => renderClient(el.dataset.k));
  }});
}}

function renderClient(k) {{
  const c = DATA.find(x => x.k === k);
  if (!c) return;
  renderSidebar(k);
  const m = document.getElementById('main');
  const handoverChip = c.handover ? `<span class="handover-chip">${{esc(c.handover)}}</span>` : '';

  const cards = [
    {{ h: 'Etapa', v: `<span class="stage-chip">${{esc(c.st)}}</span>${{handoverChip}}` }},
    {{ h: 'OB Lead',  v: esc(c.ob_lead) || '<span class="val muted">—</span>' }},
    {{ h: 'KAM',      v: esc(c.kam) || '<span class="val muted">—</span>' }},
    {{ h: 'Blocker principal', v: esc(c.blocker) || '<span class="val muted">sin bloqueos</span>' }},
    {{ h: 'Siguiente acción',  v: esc(c.next_action) || '<span class="val muted">—</span>' }},
    {{ h: 'Pain points',       v: esc(c.pain_points) || '<span class="val muted">—</span>' }},
  ];
  const cardsHTML = cards.map(x => `
    <div class="card"><h4>${{x.h}}</h4><div class="val">${{x.v}}</div></div>`).join('');

  function section(title, items) {{
    const body = items.length
      ? `<ul>${{items.map(i => `<li>${{esc(i)}}</li>`).join('')}}</ul>`
      : `<div class="empty">Sin registros</div>`;
    return `
      <div class="section">
        <header><h3>${{title}}</h3><span class="count">${{items.length}}</span></header>
        <div class="body">${{body}}</div>
      </div>`;
  }}

  const tasks = (c.tasks || []).slice().sort((a,b) => {{
    const ra = statusClass(a.s) === 'ok' ? 1 : 0;
    const rb = statusClass(b.s) === 'ok' ? 1 : 0;
    if (ra !== rb) return ra - rb;
    return (a.d || '').localeCompare(b.d || '');
  }});

  const tasksHTML = tasks.length ? `
    <div class="section">
      <header><h3>Tareas</h3><span class="count">${{tasks.length}}</span></header>
      <table class="tasks">
        <thead><tr>
          <th>Tarea</th><th>Resp.</th><th>Owner</th>
          <th>Due</th><th>Status</th>
        </tr></thead>
        <tbody>
          ${{tasks.map(t => `
            <tr>
              <td>${{esc(t.t)}}</td>
              <td>${{esc(t.r)}}</td>
              <td>${{esc(t.o)}}</td>
              <td>${{esc(t.d)}}</td>
              <td><span class="pill ${{statusClass(t.s)}}">${{esc(t.s || '—')}}</span></td>
            </tr>`).join('')}}
        </tbody>
      </table>
    </div>` : '';

  m.innerHTML = `
    <div class="client-header">
      <div>
        <h1>${{esc(c.n)}}</h1>
        <div class="sub">${{esc(c.st)}} · ${{esc(c.sx)}} · ${{c.task_count}} tareas (${{c.open_task_count}} abiertas)</div>
      </div>
      <div class="progress">
        <div class="pct">${{c.p}}%</div>
        <div class="lbl">Avance</div>
      </div>
    </div>
    <div class="cards">${{cardsHTML}}</div>
    ${{section('Next steps',  c.next_steps)}}
    ${{section('Riesgos',     c.risks)}}
    ${{section('Last week',   c.last_week)}}
    ${{tasksHTML}}
  `;
  m.scrollTop = 0;
}}

if (DATA.length) renderClient(DATA[0].k);
else document.getElementById('main').innerHTML = '<div class="empty">Sin clientes para mostrar.</div>';
</script>

</body>
</html>"""


def main():
    clients = load_clients()
    html_str = render_html(clients)
    OUT.write_text(html_str, encoding="utf-8")
    print(f"✓ {len(clients)} clientes, "
          f"{sum(c['task_count'] for c in clients)} tareas")
    print(f"✓ Wrote {OUT} ({len(html_str):,} chars)")


if __name__ == "__main__":
    main()
